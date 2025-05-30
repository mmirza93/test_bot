import streamlit as st
import pandas as pd
from openpyxl import load_workbook

file_path = "Qarabag Economic Model_2032_oil.xlsx"
result_sheet = "Result"
price_sheet = "Intro"
price_cell = "B14"
capex_sheet = "Model Inputs"
capex_items_range = "B134:B146"
capex_years_range = "E133:AK133"

def update_oil_price_and_load_data(oil_price):
    wb = load_workbook(filename=file_path)
    sht_price = wb[price_sheet]
    sht_result = wb[result_sheet]

    # Update oil price cell
    sht_price[price_cell] = oil_price

    # Save workbook to keep price updated (but no recalculation!)
    wb.save(file_path)

    # Read indicators and entities from result sheet
    indicators = [str(cell.value).strip() for cell in sht_result['B3':'B8']]
    entities = [str(cell.value).strip() for cell in sht_result['C2':'G2'][0]]
    
    # Read values range C3:G8
    values = []
    for row in sht_result['C3':'G8']:
        values.append([cell.value for cell in row])

    data = pd.DataFrame(values, index=indicators, columns=entities)
    return data

def read_current_oil_price():
    wb = load_workbook(filename=file_path, data_only=True)
    sht_price = wb[price_sheet]
    price = sht_price[price_cell].value
    return price

def read_capex_breakdown(mode="total"):
    wb = load_workbook(filename=file_path, data_only=True)
    sht = wb[capex_sheet]

    items = [cell.value.strip() if cell.value else "" for cell in sht[capex_items_range]]
    years = [cell.value for cell in sht[capex_years_range][0]]
    
    # Values range E134:AK146
    values = []
    for row in sht["E134":"AK146"]:
        values.append([cell.value if cell.value is not None else 0 for cell in row])
    
    df = pd.DataFrame(values, index=items, columns=years)
    if mode == "total":
        return df.sum(axis=1)
    elif mode == "year":
        return df.sum(axis=0)
    else:
        return None

entity_map = {
    "socar karabakh": "SOCAR Karabakh",
    "soa": "SOCAR Karabakh",
    "rsa project": "RSA Project",
    "project": "RSA Project",
    "total project": "RSA Project",
    "contractor party 2": "Contractor Party 2",
    "foreign contractor": "Contractor Party 2",
    "bp": "Contractor Party 2",
    "other contractor": "Contractor Party 2",
    "state share": "State Share",
    "state/sofaz incl. ag": "State/SOFAZ incl. AG"
}

indicator_map = {
    "capex": "CAPEX, MM USD",
    "cash flow": "Cash Flow, MM USD",
    "npv": "NPV10, MM USD",
    "npv10": "NPV10, MM USD",
    "irr": "IRR, %",
    "non-discounted payback": "Non-Discounted Payback Period",
    "discounted payback": "Discounted Payback Period"
}

def check_greeting(user_input):
    if "hello" in user_input.lower() and "ocean" in user_input.lower():
        return "Hello SUMI User, how can I help you with?"
    return None

def query_model(user_input, data):
    user_input = user_input.lower()
    greeting_response = check_greeting(user_input)
    if greeting_response:
        return greeting_response
    if "what oil price" in user_input and "used" in user_input:
        price = read_current_oil_price()
        return f"The current oil price used in the model is {price} USD/bbl."
    if "capex" in user_input and "breakdown" in user_input:
        return "Would you like the CAPEX breakdown by 'life of field' (total) or 'year by year'?"
    found_entity = next((entity_map[e] for e in entity_map if e in user_input), None)
    found_indicator = next((indicator_map[i] for i in indicator_map if i in user_input), None)
    if not found_entity and "capex" not in user_input:
        return "Sorry, I couldn't find which party you're asking about."
    if not found_indicator and "capex" not in user_input:
        return "Sorry, I couldn't understand which metric you're referring to."
    try:
        value = data.loc[found_indicator, found_entity]
        return f"{found_entity}'s {found_indicator} is {value}"
    except KeyError:
        return "Data not found for the given query."

st.title("Qarabag Economic Model Bot")

data = update_oil_price_and_load_data(70)  # default oil price

user_input = st.text_input("Ask me a question:")

if user_input:
    lowered = user_input.lower()

    if ("oil price" in lowered) and any(cmd in lowered for cmd in ["update", "set", "change"]):
        new_price = None
        for w in user_input.split():
            try:
                val = float(w)
                if val > 0:
                    new_price = val
                    break
            except:
                continue
        if new_price is not None:
            try:
                data = update_oil_price_and_load_data(new_price)
                st.success(f"Oil price updated to {new_price} USD/bbl.")
            except Exception as e:
                st.error(f"Failed to update oil price: {e}")
        else:
            st.warning("Sorry, I couldn't find a valid oil price in your request.")

    elif "life of field" in lowered or "total" in lowered:
        capex_totals = read_capex_breakdown(mode="total")
        st.write("CAPEX Breakdown - Life of Field (Total):")
        st.dataframe(capex_totals)

    elif "year by year" in lowered:
        capex_years = read_capex_breakdown(mode="year")
        st.write("CAPEX Breakdown - Year by Year:")
        st.dataframe(capex_years)

    else:
        response = query_model(user_input, data)
        st.write(response)
